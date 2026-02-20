from __future__ import annotations
from dataclasses import dataclass, asdict
from pathlib import Path
from datetime import date
from openpyxl import Workbook, load_workbook

# Normalized columns (based on your existing file, with fix for 'относно.1')
COLUMNS = [
    "Дата на получаване",
    "Вид документ",
    "Входящ номер",
    "Фактура №",
    "Наименование на фирмата / подател",
    "Относно",
    "Сума",
    "Валута",
    "Статус",
    "Бележки",
]

@dataclass(frozen=True)
class RegistryRow:
    received_date: str
    doc_type: str
    incoming_no: str | None
    invoice_no: str | None
    party: str | None
    subject: str | None
    amount: str | None
    currency: str | None
    status: str
    notes: str | None

    def to_cells(self) -> list:
        return [
            self.received_date,
            self.doc_type,
            self.incoming_no or "",
            self.invoice_no or "",
            self.party or "",
            self.subject or "",
            self.amount or "",
            self.currency or "",
            self.status,
            self.notes or "",
        ]

def ensure_workbook(path: Path) -> None:
    if path.exists():
        wb = load_workbook(str(path))
        ws = wb.active
        # If empty, set headers
        if ws.max_row == 1 and all((ws.cell(1, i+1).value is None) for i in range(len(COLUMNS))):
            for i, col in enumerate(COLUMNS, start=1):
                ws.cell(1, i).value = col
            wb.save(str(path))
        else:
            # Ensure header row matches (soft check)
            headers = [ws.cell(1, i+1).value for i in range(len(COLUMNS))]
            if headers[:len(COLUMNS)] != COLUMNS:
                # Do not overwrite; caller can decide
                pass
        return

    wb = Workbook()
    ws = wb.active
    for i, col in enumerate(COLUMNS, start=1):
        ws.cell(1, i).value = col
    wb.save(str(path))

def append_row(path: Path, row: RegistryRow) -> None:
    ensure_workbook(path)
    wb = load_workbook(str(path))
    ws = wb.active
    ws.append(row.to_cells())
    wb.save(str(path))
